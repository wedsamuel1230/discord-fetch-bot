#!/usr/bin/env python3
"""
Simple file conversion utility (CSV/XLSX, images, video via ffmpeg).

Usage examples:
  # File to file conversion
  uv run --no-project scripts/convert.py --input data.csv --to xlsx --output data.xlsx
  
  # Base64 input to file output
  uv run --no-project scripts/convert.py --input-type base64 --input <base64_data> --input-format csv --to xlsx --output data.xlsx
  
  # URL input to base64 output
  uv run --no-project scripts/convert.py --input-type url --input https://example.com/data.csv --input-format csv --to xlsx --output-type base64
  
  # Image conversion
  uv run --no-project scripts/convert.py --input image.png --to jpg --output image.jpg
"""

from __future__ import annotations

import argparse
import base64
import csv
import os
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse

import requests


IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tiff"}
VIDEO_EXTS = {".mp4", ".mov", ".mkv", ".webm", ".avi"}
AUDIO_EXTS = {".mp3", ".wav", ".flac", ".aac", ".ogg", ".m4a"}
CAD_EXTS = {".dxf"}
SUBTITLE_EXTS = {".srt", ".vtt", ".ass", ".ssa"}


def read_input_to_bytes(input_type: str, input_value: str) -> bytes:
    """
    Read input from various sources to bytes.

    Args:
        input_type: "file_path", "base64", or "url"
        input_value: The value corresponding to the input_type

    Returns:
        bytes: The input data as bytes

    Raises:
        ValueError: If input_type is invalid or input_value is malformed
        RuntimeError: For network or file errors
    """
    if input_type == "file_path":
        if not os.path.isfile(input_value):
            raise ValueError(f"Input file not found: {input_value}")
        try:
            with open(input_value, "rb") as f:
                return f.read()
        except OSError as e:
            raise RuntimeError(f"Failed to read file {input_value}: {e}") from e

    elif input_type == "base64":
        try:
            return base64.b64decode(input_value)
        except Exception as e:
            raise ValueError(f"Invalid base64 input: {e}") from e

    elif input_type == "url":
        try:
            # Validate URL
            parsed = urlparse(input_value)
            if not parsed.scheme or not parsed.netloc:
                raise ValueError(f"Invalid URL: {input_value}")

            # Download with streaming for large files
            response = requests.get(input_value, stream=True, timeout=30)
            response.raise_for_status()

            # For large files, stream to a temporary file first
            content_length = response.headers.get('content-length')
            if content_length and int(content_length) > 100 * 1024 * 1024:  # 100MB
                with tempfile.NamedTemporaryFile(delete=False) as tmp:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            tmp.write(chunk)
                    tmp.flush()
                    with open(tmp.name, "rb") as f:
                        data = f.read()
                    os.unlink(tmp.name)
                    return data
            else:
                return response.content

        except requests.RequestException as e:
            raise RuntimeError(f"Failed to download from URL {input_value}: {e}") from e
        except OSError as e:
            raise RuntimeError(f"Failed to handle downloaded content: {e}") from e

    else:
        raise ValueError(f"Unsupported input_type: {input_type}. Must be 'file_path', 'base64', or 'url'")


def write_output_from_bytes(output_type: str, output_value: str, data: bytes) -> str:
    """
    Write output data to various destinations.

    Args:
        output_type: "file_path" or "base64"
        output_value: For "file_path", the file path; for "base64", ignored
        data: The data to write

    Returns:
        str: For "file_path", the actual path written; for "base64", the base64 string

    Raises:
        ValueError: If output_type is invalid
        RuntimeError: For file write errors
    """
    if output_type == "file_path":
        try:
            # Create directory if it doesn't exist
            os.makedirs(os.path.dirname(output_value), exist_ok=True)
            with open(output_value, "wb") as f:
                f.write(data)
            return output_value
        except OSError as e:
            raise RuntimeError(f"Failed to write to file {output_value}: {e}") from e

    elif output_type == "base64":
        try:
            return base64.b64encode(data).decode('ascii')
        except Exception as e:
            raise RuntimeError(f"Failed to encode data to base64: {e}") from e

    else:
        raise ValueError(f"Unsupported output_type: {output_type}. Must be 'file_path' or 'base64'")


def convert_with_io_handling(
    input_type: str,
    input_value: str,
    output_type: str,
    output_value: str,
    target_format: str,
    input_format: Optional[str] = None,
) -> str:
    """
    Convert file with input/output handling.

    Args:
        input_type: "file_path", "base64", or "url"
        input_value: Input value (path, base64 string, or URL)
        output_type: "file_path" or "base64"
        output_value: Output value (path for file_path, ignored for base64)
        target_format: Target format extension (e.g., "xlsx", "jpg")
        input_format: Input format extension (e.g., "csv", "png") - required for base64/url

    Returns:
        str: Output result (file path or base64 string)

    Raises:
        ValueError: For invalid inputs
        RuntimeError: For conversion errors
    """
    # Read input to bytes
    input_data = read_input_to_bytes(input_type, input_value)

    # Determine input extension
    if input_type == "file_path":
        input_ext = os.path.splitext(input_value)[1].lower()
    elif input_format:
        input_ext = f".{input_format.lower().lstrip('.')}"
    else:
        raise ValueError("input_format required for base64/url inputs")

    # Create temporary input file
    with tempfile.NamedTemporaryFile(delete=False, suffix=input_ext) as tmp_input:
        tmp_input.write(input_data)
        tmp_input_path = tmp_input.name

    try:
        # Determine output path
        if output_type == "file_path":
            if not output_value:
                raise ValueError("output_value required for file_path output")
            output_path = output_value
            if not os.path.splitext(output_path)[1]:
                output_path = f"{output_path}.{target_format.lstrip('.')}"
        else:  # base64
            output_path = None

        # Perform conversion
        if output_path:
            # Direct conversion to output file
            _convert_bytes_to_file(tmp_input_path, input_ext, output_path, target_format)
            return output_path
        else:
            # Convert to temporary output file, then read to bytes
            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{target_format}") as tmp_output:
                tmp_output_path = tmp_output.name

            try:
                _convert_bytes_to_file(tmp_input_path, input_ext, tmp_output_path, target_format)
                with open(tmp_output_path, "rb") as f:
                    output_data = f.read()
                return write_output_from_bytes("base64", "", output_data)
            finally:
                if os.path.exists(tmp_output_path):
                    os.unlink(tmp_output_path)

    finally:
        if os.path.exists(tmp_input_path):
            os.unlink(tmp_input_path)


def _convert_bytes_to_file(input_path: str, input_ext: str, output_path: str, target_format: Optional[str]) -> None:
    """
    Internal function to convert using existing logic with file paths.
    """
    # Detect conversion type
    conversion = detect_conversion(input_path, output_path, target_format)

    # Perform conversion
    if conversion == "csv-to-xlsx":
        convert_csv_to_xlsx(input_path, output_path)
    elif conversion == "xlsx-to-csv":
        convert_xlsx_to_csv(input_path, output_path)
    elif conversion == "image-convert":
        convert_image(input_path, output_path)
    elif conversion == "image-to-svg":
        convert_image_to_svg(input_path, output_path)
    elif conversion == "video-convert":
        convert_video(input_path, output_path)
    elif conversion == "ocr-text":
        convert_image_to_text(input_path, output_path)
    elif conversion == "gif-optimize":
        optimize_gif(input_path, output_path)
    elif conversion == "audio-convert":
        convert_audio(input_path, output_path)
    elif conversion == "cad-convert":
        convert_cad(input_path, output_path)
    elif conversion == "subtitle-convert":
        convert_subtitle(input_path, output_path)
    else:
        raise ValueError(f"Unsupported conversion: {conversion}")


def detect_conversion(input_path: str, output_path: Optional[str], target_format: Optional[str]) -> str:
    input_ext = os.path.splitext(input_path)[1].lower()
    output_ext = None
    if output_path:
        output_ext = os.path.splitext(output_path)[1].lower()
        if not output_ext and target_format:
            output_ext = f".{target_format.lower().lstrip('.')}"
    elif target_format:
        output_ext = f".{target_format.lower().lstrip('.')}"

    if not output_ext:
        raise ValueError("Output format is required")

    if input_ext == ".csv" and output_ext == ".xlsx":
        return "csv-to-xlsx"
    if input_ext == ".xlsx" and output_ext == ".csv":
        return "xlsx-to-csv"
    if input_ext in IMAGE_EXTS and output_ext in IMAGE_EXTS:
        return "image-convert"
    if input_ext in IMAGE_EXTS and output_ext == ".svg":
        return "image-to-svg"
    if input_ext in VIDEO_EXTS and output_ext in VIDEO_EXTS:
        return "video-convert"
    if input_ext in IMAGE_EXTS and output_ext == ".txt":
        return "ocr-text"
    if input_ext == ".gif" and output_ext == ".gif":
        return "gif-optimize"
    if input_ext in AUDIO_EXTS and output_ext in AUDIO_EXTS:
        return "audio-convert"
    if input_ext in CAD_EXTS and output_ext in {".svg", ".dxf"}:  # Assuming conversion to SVG or another DXF
        return "cad-convert"
    if input_ext in SUBTITLE_EXTS and output_ext in SUBTITLE_EXTS:
        return "subtitle-convert"

    raise ValueError(f"Unsupported conversion: {input_ext} -> {output_ext}")


def validate_output_target(output_path: Optional[str], target_format: Optional[str]) -> None:
    if output_path and target_format:
        output_ext = os.path.splitext(output_path)[1].lower().lstrip(".")
        target_ext = target_format.lower().lstrip(".")
        if output_ext and output_ext != target_ext:
            raise ValueError("--output extension conflicts with --to")


def resolve_output_path(
    input_path: str,
    output_path: Optional[str],
    target_format: Optional[str],
) -> str:
    if output_path:
        ext = os.path.splitext(output_path)[1]
        if not ext and target_format:
            return f"{output_path}.{target_format.lstrip('.')}"
        return output_path

    if not target_format:
        raise ValueError("Output format is required")

    base, _ = os.path.splitext(input_path)
    return f"{base}.{target_format.lstrip('.')}"


def convert_csv_to_xlsx(input_path: str, output_path: str) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for CSV -> XLSX") from exc

    workbook = Workbook()
    sheet = workbook.active

    with open(input_path, newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        for row in reader:
            sheet.append(row)

    workbook.save(output_path)


def convert_xlsx_to_csv(input_path: str, output_path: str) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX -> CSV") from exc

    workbook = load_workbook(input_path, data_only=True)
    sheet = workbook.active

    with open(output_path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(["" if cell is None else cell for cell in row])


def convert_image(input_path: str, output_path: str) -> None:
    try:
        from PIL import Image
    except ImportError as exc:
        raise RuntimeError("Pillow is required for image conversion") from exc

    with Image.open(input_path) as img:
        output_ext = os.path.splitext(output_path)[1].lower()
        if output_ext in {".jpg", ".jpeg"} and img.mode in {"RGBA", "LA", "P"}:
            img = img.convert("RGB")
        img.save(output_path)


def convert_image_to_svg(input_path: str, output_path: str) -> None:
    """
    Convert image to SVG using potrace for high-quality vectorization.
    Uses Pillow to preprocess the image and potrace command-line tool for tracing.
    """
    try:
        from PIL import Image
    except ImportError as exc:
        raise RuntimeError("Pillow is required for image to SVG conversion") from exc

    # Check if potrace is available
    potrace_path = Path(__file__).parent.parent.parent.parent / "potrace" / "potrace-1.16.win64" / "potrace.exe"
    if not potrace_path.exists():
        # Try relative path from current working directory
        alt_path = Path("../../../potrace/potrace-1.16.win64/potrace.exe")
        if alt_path.exists():
            potrace_path = alt_path
        else:
            raise RuntimeError("potrace executable not found. Please install potrace system-wide or ensure it's in the expected location.")

    try:
        with Image.open(input_path) as img:
            # Handle transparent backgrounds
            if img.mode == 'RGBA':
                # Create a white background
                background = Image.new('RGBA', img.size, (255, 255, 255, 255))
                # Composite the image onto the white background
                img = Image.alpha_composite(background, img)
                # Convert to RGB
                img = img.convert('RGB')

            # Convert to grayscale
            img = img.convert("L")

            # Scale up for better tracing (2x as suggested)
            width, height = img.size
            img = img.resize((width * 2, height * 2), Image.Resampling.LANCZOS)

            # Apply threshold to create binary image (adjustable threshold)
            threshold = 180  # Starting point as suggested
            bw = img.point(lambda p: 255 if p > threshold else 0)
            bw = bw.convert("1")  # 1-bit black and white

            # Save as PBM (Portable Bitmap) format
            with tempfile.NamedTemporaryFile(suffix='.pbm', delete=False) as temp_pbm:
                pbm_path = temp_pbm.name
                bw.save(pbm_path, format='PPM')  # PBM is a type of PPM

            try:
                # Call potrace to convert PBM to SVG
                command = [
                    str(potrace_path),
                    pbm_path,
                    "-s",  # SVG output
                    "-o", output_path
                ]

                result = subprocess.run(command, check=True, capture_output=True, text=True)
                if result.returncode != 0:
                    raise RuntimeError(f"potrace failed: {result.stderr}")

            finally:
                # Clean up temporary PBM file
                Path(pbm_path).unlink(missing_ok=True)

    except Exception as e:
        raise RuntimeError(f"Image to SVG conversion failed: {e}") from e


def convert_video(input_path: str, output_path: str) -> None:
    if shutil.which("ffmpeg") is None:
        raise RuntimeError("ffmpeg is required for video conversion")
    command = ["ffmpeg", "-y", "-i", input_path, output_path]
    completed = subprocess_run(command)
    if completed != 0:
        raise RuntimeError("ffmpeg conversion failed")


def convert_image_to_text(input_path: str, output_path: str) -> None:
    try:
        import pytesseract
        from PIL import Image
    except ImportError as exc:
        raise RuntimeError("pytesseract and Pillow are required for OCR text extraction") from exc

    try:
        img = Image.open(input_path)
        text = pytesseract.image_to_string(img)
        Path(output_path).write_text(text, encoding='utf-8')
    except Exception as e:
        raise RuntimeError(f"OCR extraction failed: {e}") from e


def optimize_gif(input_path: str, output_path: str) -> None:
    if shutil.which("gifsicle") is None:
        raise RuntimeError("gifsicle is required for GIF optimization")
    command = ["gifsicle", "--optimize=3", "--output", output_path, input_path]
    completed = subprocess_run(command)
    if completed != 0:
        raise RuntimeError("gifsicle optimization failed")


def convert_audio(input_path: str, output_path: str) -> None:
    try:
        import ffmpeg
    except ImportError as exc:
        raise RuntimeError("ffmpeg-python is required for audio conversion") from exc

    try:
        stream = ffmpeg.input(input_path)
        stream = ffmpeg.output(stream, output_path)
        ffmpeg.run(stream, overwrite_output=True, quiet=True)
    except ffmpeg.Error as e:
        raise RuntimeError(f"Audio conversion failed: {e}") from e


def convert_cad(input_path: str, output_path: str) -> None:
    try:
        import ezdxf
        from ezdxf.addons.drawing import Frontend, RenderContext, layout, svg
        from ezdxf import bbox
    except ImportError as exc:
        raise RuntimeError("ezdxf is required for CAD file conversion") from exc

    try:
        doc = ezdxf.readfile(input_path)
        output_ext = Path(output_path).suffix.lower()
        if output_ext == ".svg":
            # Use ezdxf's proper SVG rendering which handles all entity types including SPLINE
            msp = doc.modelspace()
            
            # Calculate bounding box for proper viewport setup
            try:
                extents = bbox.extents(msp)
                if extents:
                    width = abs(extents.extmax.x - extents.extmin.x)
                    height = abs(extents.extmax.y - extents.extmin.y)
                    # Add some padding
                    width *= 1.1
                    height *= 1.1
                    page = layout.Page(width, height)
                else:
                    # Default page size if no extents
                    page = layout.Page(1000, 1000)
            except:
                # Fallback if bbox calculation fails
                page = layout.Page(1000, 1000)
            
            backend = svg.SVGBackend()
            frontend = Frontend(RenderContext(doc), backend)
            frontend.draw_layout(msp)
            
            # Get the SVG content and write to file
            svg_content = backend.get_string(page)
            
            # Post-process SVG to add transparent background and black outlines
            if '<svg' in svg_content:
                # Insert style to make background transparent and outlines black
                style_insert = '<defs><style>svg { background-color: transparent; } rect[fill] { fill: transparent !important; } .C1 { stroke: #000000 !important; }</style></defs>'
                # Use regex to properly insert defs/style after the opening <svg> tag
                svg_content = re.sub(r'(<svg[^>]*>)', rf'\1{style_insert}', svg_content, count=1)
            
            Path(output_path).write_text(svg_content, encoding='utf-8')
        elif output_ext == ".dxf":
            # Just copy if same format, or implement conversion
            doc.saveas(output_path)
        else:
            raise ValueError(f"Unsupported CAD output format: {output_ext}")
    except Exception as e:
        raise RuntimeError(f"CAD conversion failed: {e}") from e


def convert_subtitle(input_path: str, output_path: str) -> None:
    try:
        import pysrt
    except ImportError as exc:
        raise RuntimeError("pysrt is required for subtitle conversion") from exc

    try:
        input_ext = Path(input_path).suffix.lower()
        output_ext = Path(output_path).suffix.lower()
        
        if input_ext == ".srt":
            subs = pysrt.open(input_path)
            if output_ext == ".vtt":
                # Convert SRT to VTT
                vtt_content = "WEBVTT\n\n"
                for sub in subs:
                    start = sub.start.to_time()
                    end = sub.end.to_time()
                    vtt_content += f"{sub.index}\n{start} --> {end}\n{sub.text}\n\n"
                Path(output_path).write_text(vtt_content, encoding='utf-8')
            elif output_ext == ".srt":
                subs.save(output_path, encoding='utf-8')
            else:
                raise ValueError(f"Unsupported subtitle output format: {output_ext}")
        else:
            raise ValueError(f"Unsupported subtitle input format: {input_ext}")
    except Exception as e:
        raise RuntimeError(f"Subtitle conversion failed: {e}") from e


def subprocess_run(command: list[str]) -> int:
    import subprocess
    completed = subprocess.run(command, check=False)
    return completed.returncode


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert files between formats")
    parser.add_argument("--input", help="Input file path")
    parser.add_argument("--input-type", choices=["file_path", "base64", "url"], default="file_path", help="Input type (default: file_path)")
    parser.add_argument("--input-format", help="Input format extension (required for base64/url inputs)")
    parser.add_argument("--output", help="Output file path (for file_path output)")
    parser.add_argument("--output-type", choices=["file_path", "base64"], default="file_path", help="Output type (default: file_path)")
    parser.add_argument("--to", dest="target_format", required=True, help="Target format (extension)")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    
    if not args.input:
        print("Provide --input", file=sys.stderr)
        return 2

    # Validate input requirements
    if args.input_type in ["base64", "url"] and not args.input_format:
        print("--input-format required for base64/url inputs", file=sys.stderr)
        return 2

    if args.output_type == "file_path" and not args.output:
        print("--output required for file_path output", file=sys.stderr)
        return 2

    try:
        result = convert_with_io_handling(
            input_type=args.input_type,
            input_value=args.input,
            output_type=args.output_type,
            output_value=args.output or "",
            target_format=args.target_format,
            input_format=args.input_format,
        )
        
        if args.output_type == "file_path":
            print(f"Saved: {result}")
        else:
            print(result)  # base64 output
        return 0
        
    except (ValueError, RuntimeError) as exc:
        print(str(exc), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
